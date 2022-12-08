###This code filters the original timber forestry dataset to include only the species suitable for clt manufacturing

######################################################################################################
# import all required packages and functions
# from Functions import get_GM_API_Key
# from Functions import lumber_species_properties_processing
# from Functions import calculate_lumber_impact
# from Functions import calculate_distance
# from Functions import calculate_energy_impacts
from Google_Maps_Functions import calculate_distance
from Google_Maps_Functions import get_GM_API_Key
from Google_Maps_Functions import haversine
import googlemaps
import gmaps
import gmaps.datasets
import json
import requests
import pandas as pd
import xlsxwriter
import warnings
warnings.filterwarnings('ignore')
import math as math
from openpyxl import load_workbook
import xlsxwriter
import numpy as np
#import matplotlib
#import matplotlib.pypl

# from itertools import tee ##this line may not be needed
###import api_key from text file and configure the key
apikey_text = get_GM_API_Key()
#print(apikey_text)
gmaps.configure(api_key=apikey_text)

#############

states = pd.read_excel(".../GitHub/CLT-LCA-Tool/Data Sheets/CLT_Timber_types.xlsx", sheet_name = 'CLT_Suit_Species')
#states_list = states["State"].tolist()
sawmills_dataset = pd.read_excel(".../CLT-LCA-Tool/Mill Datasets/Sawmills_Dataset.xlsx")
state_to_region = pd.read_excel(".../CLT-LCA-Tool/Mill Datasets/State_to_Region.xlsx",sheet_name="Sheet1")

regional_max_distance = pd.read_excel(".../CLT-LCA-Tool/Mill Datasets/State_to_Region.xlsx",sheet_name="Sheet2")

##NH,NJ,NY,NV,OH,SC,TN,TX,VA,WI,PA
#state_list_pnw = ['WA', 'OR', 'CA']
state = 'WI'
region = state_to_region[state_to_region["State"] == state]["State_Region"].item()
print(region)
max_forest_mill_distance = regional_max_distance[regional_max_distance["Region"] == region]["Max_Distance"].item()
print(max_forest_mill_distance)

state_forest_data = pd.read_excel(".../CLT-LCA-Tool/Mill Datasets/Forestry_Data_GIS/Forestry_Data_GIS_by_State/G5km_OutputSummary_"+state+".xlsx")
timber_species_in_state = state_forest_data["Forest_Type"].unique().tolist()
print(timber_species_in_state)
sawmills_state = sawmills_dataset[sawmills_dataset["STATE"] == state]
print(sawmills_state)
#print(timber_species_in_state)
# lat1 = 48.01
# long1 = -121.53
# lat1 = 45.639
# long1 = -122.624
# cord1 = str(lat1) +', '+ str(long1)
#
# lat2 = 46.40
# long2 = -117.07
# cord2 = str(lat2) +', '+ str(long2)
#
# distance_f_m = calculate_distance(cord1, cord2, apikey_text)
# print(distance_f_m)

############################
state_timber_sawmill_dataset = pd.DataFrame(columns=['Timber_Type', 'Mill_ID_U', 'LON', 'LAT', 'STATE', 'PRECISE_TY', 'Total Area', 'Available', 'Haver_ltt', 'Nan Count','Nan Area'])
book = load_workbook('.../CLT-LCA-Tool/Mill Datasets/state_timber_sawmill_dataset_'+state+'.xlsx')
writer = pd.ExcelWriter('.../CLT-LCA-Tool/Mill Datasets/state_timber_sawmill_dataset_'+state+'.xlsx', engine ='openpyxl')
writer.book = book
#workbook = xlsxwriter.Workbook('C:/Users/SATNOORK/Documents/GitHub/CLT-LCA-Tool/Mill Datasets/state_timber_sawmill_dataset_'+state+'.xlsx')
#worksheet = workbook.add_worksheet()
writer.save()
xyz = True
for timber_species in timber_species_in_state:
    print(timber_species)
    #if not timber_species in ('Baldcypress/Water Tupelo','Eastern Redcedar','Eastern Redcedar/Hardwood','Eastern White Pine/Northern Red','Juniper Woodland','Loblolly Pine','Loblolly Pine/Hardwood','Longleaf Pine','Longleaf Pine/Oak','Misc. Western Softwoods','Pinyon Juniper Woodland','Ponderosa Pine','Rocky Mountain Juniper','Shortleaf Pine','Shortleaf Pine/Oak','Slash Pine','Slash Pine/Hardwood','White Fir'):
    #xyz == True
    if timber_species not in ['Eastern Hemlock','Eastern Redcedar/Hardwood','Eastern Redcedar','Eastern White Pine','Eastern White Pine/Northern Red','Jack Pine','Northern White-cedar','Red Pine','Scotch Pine','Tamarack','White Spruce']:
            #timber_species in ['Douglas-fir', 'Eastern Hemlock','Eastern Redcedar/Hardwood']:

        timber_state_forest_data = state_forest_data[state_forest_data["Forest_Type"] == timber_species]
        timber_state_forest_data = timber_state_forest_data.sort_values("Area_km2", ascending = False)
        print(sawmills_state)
        #print(timber_state_forest_data["Area_km2"])
        #print(timber_species)
        for index, row in sawmills_state.iterrows():
            mill_coordinates = str(row['LAT']) + ', ' + str(row['LON'])
            mill_lat = row['LAT']
            mill_long = row['LON']
            print(row["Mill_ID_U"])
            total_area = 0
            nan_area = 0
            count_mill = 0
            index_count = 0
            nan_count = 0
            haversine_less_than_threshold = 0
            #haversine_area_wtd_avg = 0
            #print(timber_state_forest_data)
            for index_f, row_f in timber_state_forest_data.iterrows():
                #print("inside")
                forest_coordinates = str(row_f['G5km_Lat']) + ', ' + str(row_f['G5km_Lon'])
                for_lat = row_f['G5km_Lat']
                for_long = row_f['G5km_Lon']
                #print(forest_coordinates)
                distance_haversine = haversine(mill_long, mill_lat, for_long, for_lat)
                #print(distance_haversine)
                #print(2)
                if distance_haversine <= max_forest_mill_distance* 1.05:
                    #print("heversine less than 250")
                    distance_f_m = calculate_distance(forest_coordinates, mill_coordinates, apikey_text)
                    #print(distance_f_m)
                    haversine_less_than_threshold = haversine_less_than_threshold + 1
                    #print(index_f)
                    #print(distance_haversine)
                    #previous_total_area = Sawmills_WA["Total Area"].iloc[index_count]
                    #print(3)
                    if math.isnan(distance_f_m) == False:
                        if distance_f_m <= max_forest_mill_distance:
                            total_area = total_area + row_f["Area_km2"]
                            print("yeyyy")
                            #print(total_area)
                        ####
                        if total_area >= 5:
                            print('break')
                            break
                        ####
                    ####
                    if math.isnan(distance_f_m) == True:
                        nan_count = nan_count + 1
                        nan_area = nan_area + row_f["Area_km2"]
                    ####
                ####
            ####
            append_to_state_timber_sawmill_dataset = pd.DataFrame(columns=['Timber_Type', 'Mill_ID_U', 'LON', 'LAT', 'STATE', 'PRECISE_TY', 'Total Area', 'Available', 'Haver_ltt', 'Nan Count','Nan Area'])
            append_to_state_timber_sawmill_dataset = append_to_state_timber_sawmill_dataset.append(pd.Series(), ignore_index=True)

            append_to_state_timber_sawmill_dataset.iloc[0]["Timber_Type"] = timber_species
            #state_timber_sawmill_dataset = state_timber_sawmill_dataset.append(row)
            append_to_state_timber_sawmill_dataset.iloc[0]["Mill_ID_U"] = row["Mill_ID_U"]
            append_to_state_timber_sawmill_dataset.iloc[0]["LON"] = row["LON"]
            append_to_state_timber_sawmill_dataset.iloc[0]["LAT"] = row["LAT"]
            append_to_state_timber_sawmill_dataset.iloc[0]["STATE"] = row["STATE"]
            append_to_state_timber_sawmill_dataset.iloc[0]["PRECISE_TY"] = row["PRECISE_TY"]
            append_to_state_timber_sawmill_dataset.iloc[0]["Total Area"] = total_area
            append_to_state_timber_sawmill_dataset.iloc[0]["Nan Count"] = nan_count
            append_to_state_timber_sawmill_dataset.iloc[0]["Haver_ltt"] = haversine_less_than_threshold
            append_to_state_timber_sawmill_dataset.iloc[0]["Nan Area"] = nan_area
            if total_area >= 5:
                append_to_state_timber_sawmill_dataset.loc[index_count]["Available"] = 'Y'
            else:
                append_to_state_timber_sawmill_dataset.loc[index_count]["Available"] = 'N'
            index_count = index_count + 1
            #print(append_to_state_timber_sawmill_dataset.iloc[0])
            #print(index_f)
            state_timber_sawmill_dataset = state_timber_sawmill_dataset.append(append_to_state_timber_sawmill_dataset)
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            #print(state_timber_sawmill_dataset)
            state_timber_sawmill_dataset.to_excel(writer, sheet_name=state)
            #worksheet.write(state_timber_sawmill_dataset)
            writer.save()
        ####

####
#print(state_timber_sawmill_dataset)
writer.close()
